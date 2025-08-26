import openpyxl
from config import *
from common.Md5_Encrypt import eg

class FileReader:
    @staticmethod
    def read_excel(file_path=EXCEL_URL, sheet_name=SHEET_NAME) -> list[dict] | None:
        """
        读取Excel文件，第二行为表头(key)，第三行开始为数据(value)
        返回列表字典格式：[{key1:value1, key2:value2}, ...]
        """
        workbook = None
        try:
            # 1. 加载工作簿（只读模式，读取公式结果）
            workbook = openpyxl.load_workbook(
                filename=file_path,
                read_only=True,
                data_only=True
            )

            # 2. 检查工作表是否存在
            if sheet_name not in workbook.sheetnames:
                available_sheets = ", ".join(workbook.sheetnames)
                print(f"警告：工作表 '{sheet_name}' 不存在！可选工作表：{available_sheets}")
                return None

            # 3. 获取工作表对象
            sheet = workbook[sheet_name]

            # 4. 读取第二行作为表头（key）
            header = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]

            # 5. 从第三行开始读取数据
            data = []
            for row in sheet.iter_rows(min_row=3, values_only=True):
                if any(cell is not None for cell in row):  # 跳过空行
                    new_data = dict(zip(header, row))
                    if new_data.get("is_true") is True:
                        data.append(new_data)
            return data

        except FileNotFoundError:
            print(f"错误：文件 '{file_path}' 不存在！")
            return None
        except Exception as e:
            print(f"读取Excel失败：{str(e)}")
            return None
        finally:
            if 'workbook' in locals():  # 确保工作簿被关闭
                workbook.close()


    @staticmethod
    def write_data_excel(file_path = EXCEL_URL, sheet_name = SHEET_NAME, row=None, column=None, value=None):
        try:
            workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.create_sheet(sheet_name)

        worksheet.cell(row=row, column=column, value=value)
        workbook.save(file_path)

    @staticmethod
    def encrypt_data_aes(data):
        new_data = {}
        for key in data:
            if key[0] == "@":
                new_data[key[1:]] = eg.encrypt(data[key])
            else:
                new_data[key] = data[key]
        return new_data



